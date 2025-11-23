from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Avg, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from .models import *
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import json


# ==================== UTILIDADES ====================

def es_estudiante(user):
    return user.is_authenticated and user.rol == 'estudiante'

def es_profesor(user):
    return user.is_authenticated and user.rol == 'profesor'

def es_administrador(user):
    return user.is_authenticated and user.rol == 'administrador'

def registrar_actividad(request, accion, modelo, objeto_id, descripcion):
    """Registra actividad en el log"""
    LogActividad.objects.create(
        usuario=request.user,
        accion=accion,
        modelo=modelo,
        objeto_id=objeto_id,
        descripcion=descripcion,
        ip_address=request.META.get('REMOTE_ADDR')
    )


# ==================== AUTENTICACIÓN ====================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Credenciales incorrectas')
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    """Vista de logout"""
    registrar_actividad(request, 'consultar', 'Usuario', request.user.id, 'Cierre de sesión')
    messages.info(request, 'Sesión cerrada correctamente')
    logout(request)
    return redirect('login')


# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    """Dashboard principal según rol del usuario"""
    user = request.user
    
    # Obtener notificaciones no leídas
    notificaciones = user.notificaciones.filter(leida=False).order_by('-fecha_creacion')[:5]
    
    context = {
        'notificaciones': notificaciones,
        'notificaciones_count': user.notificaciones.filter(leida=False).count(),
    }
    
    if user.rol == 'estudiante':
        estudiante = user.perfil_estudiante
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        
        # Inscripciones del periodo actual
        inscripciones = estudiante.inscripciones.filter(curso__periodo=periodo_actual)
        
        # Calcular promedio general
        promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
        promedio_general = sum(promedios) / len(promedios) if promedios else 0.0
        
        # Materias aprobadas/reprobadas
        materias_aprobadas = sum(1 for p in promedios if p >= 3.0)
        materias_reprobadas = sum(1 for p in promedios if p < 3.0)
        
        # Total de créditos
        total_creditos = sum(insc.curso.materia.creditos for insc in inscripciones)
        
        context.update({
            'estudiante': estudiante,
            'inscripciones': inscripciones,
            'promedio_general': round(promedio_general, 2),
            'periodo_actual': periodo_actual,
            'total_materias': inscripciones.count(),
            'materias_aprobadas': materias_aprobadas,
            'materias_reprobadas': materias_reprobadas,
            'total_creditos': total_creditos,
        })
        return render(request, 'estudiante/dashboard.html', context)
    
    elif user.rol == 'profesor':
        profesor = user.perfil_profesor
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        
        # Cursos del profesor en el periodo actual
        cursos = profesor.cursos.filter(periodo=periodo_actual)
        
        # Estadísticas
        total_estudiantes = sum(curso.estudiantes_inscritos() for curso in cursos)
        calificaciones_pendientes = 0
        
        for curso in cursos:
            inscripciones = curso.inscripciones.count()
            tipos_eval = ConfiguracionEvaluacion.objects.filter(curso=curso).count()
            calificaciones_registradas = Calificacion.objects.filter(
                inscripcion__curso=curso
            ).count()
            calificaciones_esperadas = inscripciones * tipos_eval
            calificaciones_pendientes += max(0, calificaciones_esperadas - calificaciones_registradas)
        
        context.update({
            'profesor': profesor,
            'cursos': cursos,
            'periodo_actual': periodo_actual,
            'total_estudiantes': total_estudiantes,
            'calificaciones_pendientes': calificaciones_pendientes,
        })
        return render(request, 'profesor/dashboard.html', context)
    
    elif user.rol == 'administrador':
        administrador = user.perfil_administrador
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        
        # Estadísticas generales
        total_estudiantes = Estudiante.objects.filter(estado='activo').count()
        total_profesores = Profesor.objects.count()
        total_cursos = Curso.objects.filter(periodo=periodo_actual).count()
        
        # Promedio institucional
        cursos = Curso.objects.filter(periodo=periodo_actual)
        promedios_cursos = []
        for curso in cursos:
            inscripciones = curso.inscripciones.all()
            promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
            if promedios:
                promedios_cursos.append(sum(promedios) / len(promedios))
        
        promedio_institucional = sum(promedios_cursos) / len(promedios_cursos) if promedios_cursos else 0.0
        
        # Actividad reciente
        actividades_recientes = LogActividad.objects.all().order_by('-fecha')[:10]
        
        context.update({
            'administrador': administrador,
            'total_estudiantes': total_estudiantes,
            'total_profesores': total_profesores,
            'total_cursos': total_cursos,
            'promedio_institucional': round(promedio_institucional, 2),
            'periodo_actual': periodo_actual,
            'actividades_recientes': actividades_recientes,
        })
        return render(request, 'administrador/dashboard.html', context)
    
    return redirect('login')


# ==================== ESTUDIANTE ====================

@login_required
@user_passes_test(es_estudiante)
def mis_notas(request):
    """Vista de notas del estudiante con filtros"""
    estudiante = request.user.perfil_estudiante
    periodo_id = request.GET.get('periodo')
    
    if periodo_id:
        inscripciones = estudiante.inscripciones.filter(curso__periodo_id=periodo_id)
        periodo_actual = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        inscripciones = estudiante.inscripciones.filter(curso__periodo=periodo_actual)
    
    periodos = PeriodoAcademico.objects.all()
    
    # Calcular promedio general del periodo
    promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
    promedio_general = sum(promedios) / len(promedios) if promedios else 0.0
    
    # Preparar datos para cada inscripción
    inscripciones_data = []
    for insc in inscripciones:
        promedio = insc.calcular_promedio()
        inscripciones_data.append({
            'inscripcion': insc,
            'promedio': promedio,
            'estado': insc.estado_aprobacion(),
            'calificaciones': insc.calificaciones.all(),
        })
    
    context = {
        'inscripciones_data': inscripciones_data,
        'periodos': periodos,
        'periodo_actual': periodo_actual,
        'promedio_general': round(promedio_general, 2),
    }
    
    return render(request, 'estudiante/mis_notas.html', context)

@login_required
@user_passes_test(es_estudiante)
def detalle_materia(request, inscripcion_id):
    """Detalle completo de una materia específica"""
    inscripcion = get_object_or_404(InscripcionCurso, id=inscripcion_id, estudiante=request.user.perfil_estudiante)
    calificaciones = inscripcion.calificaciones.all().order_by('-fecha_registro')
    
    # Obtener configuración de evaluaciones
    configuraciones = ConfiguracionEvaluacion.objects.filter(curso=inscripcion.curso)
    
    context = {
        'inscripcion': inscripcion,
        'calificaciones': calificaciones,
        'promedio': inscripcion.calcular_promedio(),
        'estado': inscripcion.estado_aprobacion(),
        'configuraciones': configuraciones,
    }
    
    return render(request, 'estudiante/detalle_materia.html', context)

@login_required
@user_passes_test(es_estudiante)
def actualizar_perfil(request):
    """Actualizar datos personales del estudiante"""
    if request.method == 'POST':
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        
        # Validaciones
        if not email:
            messages.error(request, 'El correo electrónico es obligatorio')
            return redirect('actualizar_perfil')
        
        request.user.email = email
        request.user.telefono = telefono
        request.user.save()
        
        registrar_actividad(request, 'editar', 'Usuario', request.user.id, 'Actualización de perfil')
        messages.success(request, 'Perfil actualizado correctamente')
        return redirect('dashboard')
    
    return render(request, 'estudiante/actualizar_perfil.html')

@login_required
@user_passes_test(es_estudiante)
def descargar_boletin(request, periodo_id):
    """Descargar boletín de notas en PDF"""
    estudiante = request.user.perfil_estudiante
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id)
    inscripciones = estudiante.inscripciones.filter(curso__periodo=periodo)
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Encabezado
    titulo = Paragraph(f"<b>BOLETÍN DE NOTAS ACADÉMICAS</b>", styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 0.3*inch))
    
    # Información del estudiante
    info_estudiante = Paragraph(f"""
        <b>Estudiante:</b> {estudiante.usuario.get_full_name()}<br/>
        <b>Código:</b> {estudiante.codigo_estudiantil}<br/>
        <b>Programa:</b> {estudiante.programa.nombre}<br/>
        <b>Semestre:</b> {estudiante.semestre}<br/>
        <b>Periodo:</b> {periodo.nombre}<br/>
        <b>Fecha de emisión:</b> {datetime.now().strftime('%d/%m/%Y')}
    """, styles['Normal'])
    elements.append(info_estudiante)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de notas
    data = [['Código', 'Materia', 'Créditos', 'Promedio', 'Estado']]
    
    total_creditos = 0
    promedios_list = []
    
    for insc in inscripciones:
        promedio = insc.calcular_promedio() or 0.0
        estado = insc.estado_aprobacion()
        creditos = insc.curso.materia.creditos
        
        data.append([
            insc.curso.materia.codigo,
            insc.curso.materia.nombre,
            str(creditos),
            f"{promedio:.2f}",
            estado
        ])
        
        total_creditos += creditos
        if promedio > 0:
            promedios_list.append(promedio)
    
    # Calcular promedio general
    promedio_general = sum(promedios_list) / len(promedios_list) if promedios_list else 0.0
    
    table = Table(data, colWidths=[1*inch, 3*inch, 1*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    # Resumen
    elements.append(Spacer(1, 0.3*inch))
    resumen = Paragraph(f"""
        <b>RESUMEN ACADÉMICO</b><br/>
        Total de Créditos: {total_creditos}<br/>
        Promedio General del Periodo: <b>{promedio_general:.2f}</b><br/>
        Estado: <b>{'APROBADO' if promedio_general >= 3.0 else 'REPROBADO'}</b>
    """, styles['Normal'])
    elements.append(resumen)
    
    # Pie de página
    elements.append(Spacer(1, 0.5*inch))
    pie = Paragraph("""
        <i>Este es un documento oficial emitido por el Sistema de Gestión Académica<br/>
        Universidad Cooperativa de Colombia - Campus Pasto</i>
    """, styles['Normal'])
    elements.append(pie)
    
    doc.build(elements)
    buffer.seek(0)
    
    registrar_actividad(request, 'consultar', 'Boletin', periodo_id, f'Descarga de boletín - {periodo.nombre}')
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="boletin_{periodo.nombre}_{estudiante.codigo_estudiantil}.pdf"'
    return response


# ==================== PROFESOR ====================

@login_required
@user_passes_test(es_profesor)
def mis_cursos(request):
    """Lista de cursos del profesor con estadísticas"""
    profesor = request.user.perfil_profesor
    periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
    cursos = profesor.cursos.filter(periodo=periodo_actual)
    
    # Agregar estadísticas a cada curso
    cursos_data = []
    for curso in cursos:
        inscripciones = curso.inscripciones.all()
        promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
        promedio_curso = sum(promedios) / len(promedios) if promedios else 0.0
        
        cursos_data.append({
            'curso': curso,
            'total_estudiantes': inscripciones.count(),
            'promedio_curso': round(promedio_curso, 2),
        })
    
    context = {
        'cursos_data': cursos_data,
        'periodo_actual': periodo_actual,
    }
    
    return render(request, 'profesor/mis_cursos.html', context)

@login_required
@user_passes_test(es_profesor)
def estudiantes_curso(request, curso_id):
    """Lista de estudiantes de un curso con sus notas"""
    curso = get_object_or_404(Curso, id=curso_id, profesor=request.user.perfil_profesor)
    inscripciones = curso.inscripciones.all().order_by('estudiante__usuario__last_name')
    
    # Preparar datos de estudiantes con sus promedios
    estudiantes_data = []
    for insc in inscripciones:
        promedio = insc.calcular_promedio()
        estudiantes_data.append({
            'inscripcion': insc,
            'estudiante': insc.estudiante,
            'promedio': promedio,
            'estado': insc.estado_aprobacion(),
            'calificaciones': insc.calificaciones.all(),
        })
    
    # Tipos de evaluación configurados
    tipos_evaluacion = ConfiguracionEvaluacion.objects.filter(curso=curso)
    
    context = {
        'curso': curso,
        'estudiantes_data': estudiantes_data,
        'tipos_evaluacion': tipos_evaluacion,
    }
    
    return render(request, 'profesor/estudiantes_curso.html', context)

@login_required
@user_passes_test(es_profesor)
def registrar_calificacion(request, inscripcion_id):
    """Registrar o editar calificación (FUNCIONALIDAD PRINCIPAL 1)"""
    inscripcion = get_object_or_404(InscripcionCurso, id=inscripcion_id)
    
    # Verificar que el profesor pertenece al curso
    if inscripcion.curso.profesor != request.user.perfil_profesor:
        messages.error(request, 'No tiene permiso para calificar este curso')
        return redirect('mis_cursos')
    
    if request.method == 'POST':
        tipo_evaluacion_id = request.POST.get('tipo_evaluacion')
        nota = request.POST.get('nota')
        observaciones = request.POST.get('observaciones', '')
        
        # Validaciones
        try:
            nota_decimal = float(nota)
            if nota_decimal < 0.0 or nota_decimal > 5.0:
                messages.error(request, 'La nota debe estar entre 0.0 y 5.0')
                return redirect('registrar_calificacion', inscripcion_id=inscripcion_id)
        except ValueError:
            messages.error(request, 'Nota inválida')
            return redirect('registrar_calificacion', inscripcion_id=inscripcion_id)
        
        calificacion, created = Calificacion.objects.update_or_create(
            inscripcion=inscripcion,
            tipo_evaluacion_id=tipo_evaluacion_id,
            defaults={
                'nota': nota_decimal,
                'observaciones': observaciones,
                'registrada_por': request.user
            }
        )
        
        # Crear notificación al estudiante
        tipo_evaluacion = TipoEvaluacion.objects.get(id=tipo_evaluacion_id)
        Notificacion.objects.create(
            usuario=inscripcion.estudiante.usuario,
            tipo='nueva_nota' if created else 'modificacion_nota',
            titulo=f"{'Nueva nota' if created else 'Nota modificada'} en {inscripcion.curso.materia.nombre}",
            mensaje=f"Se ha {'registrado' if created else 'modificado'} tu nota de {tipo_evaluacion.nombre}: {nota_decimal}"
        )
        
        accion = 'crear' if created else 'editar'
        registrar_actividad(request, accion, 'Calificacion', calificacion.id, 
                          f"{accion.capitalize()} calificación para {inscripcion.estudiante.usuario.get_full_name()}")
        
        messages.success(request, f'Calificación {"registrada" if created else "actualizada"} correctamente')
        return redirect('estudiantes_curso', curso_id=inscripcion.curso.id)
    
    tipos_evaluacion = TipoEvaluacion.objects.all()
    calificaciones_existentes = inscripcion.calificaciones.all()
    configuraciones = ConfiguracionEvaluacion.objects.filter(curso=inscripcion.curso)
    
    context = {
        'inscripcion': inscripcion,
        'tipos_evaluacion': tipos_evaluacion,
        'calificaciones_existentes': calificaciones_existentes,
        'configuraciones': configuraciones,
    }
    
    return render(request, 'profesor/registrar_calificacion.html', context)

@login_required
@user_passes_test(es_profesor)
@require_http_methods(["POST"])
def eliminar_calificacion(request, calificacion_id):
    """Eliminar una calificación (Modal/AJAX)"""
    calificacion = get_object_or_404(Calificacion, id=calificacion_id)
    
    # Verificar permisos
    if calificacion.inscripcion.curso.profesor != request.user.perfil_profesor:
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    estudiante = calificacion.inscripcion.estudiante.usuario.get_full_name()
    tipo_eval = calificacion.tipo_evaluacion.nombre
    
    calificacion.delete()
    
    registrar_actividad(request, 'eliminar', 'Calificacion', calificacion_id, 
                       f"Eliminó calificación de {tipo_eval} para {estudiante}")
    
    return JsonResponse({'success': True, 'message': 'Calificación eliminada correctamente'})


# ==================== ADMINISTRADOR ====================

@login_required
@user_passes_test(es_administrador)
def gestion_cursos(request):
    """Gestión de cursos con filtros y búsqueda"""
    periodo_id = request.GET.get('periodo')
    busqueda = request.GET.get('q', '')
    
    if periodo_id:
        cursos = Curso.objects.filter(periodo_id=periodo_id)
        periodo_actual = PeriodoAcademico.objects.get(id=periodo_id)
    else:
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        cursos = Curso.objects.filter(periodo=periodo_actual)
    
    # Búsqueda
    if busqueda:
        cursos = cursos.filter(
            Q(materia__nombre__icontains=busqueda) |
            Q(materia__codigo__icontains=busqueda) |
            Q(profesor__usuario__first_name__icontains=busqueda) |
            Q(profesor__usuario__last_name__icontains=busqueda)
        )
    
    cursos = cursos.order_by('materia__nombre')
    
    # Paginación
    paginator = Paginator(cursos, 10)
    page_number = request.GET.get('page')
    cursos_page = paginator.get_page(page_number)
    
    periodos = PeriodoAcademico.objects.all()
    
    context = {
        'cursos': cursos_page,
        'periodos': periodos,
        'periodo_actual': periodo_actual,
        'busqueda': busqueda,
    }
    
    return render(request, 'administrador/gestion_cursos.html', context)

@login_required
@user_passes_test(es_administrador)
def generar_reporte(request):
    """Generar reportes académicos (FUNCIONALIDAD PRINCIPAL 3)"""
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte')
        formato = request.POST.get('formato')  # pdf o excel
        periodo_id = request.POST.get('periodo')
        programa_id = request.POST.get('programa', None)
        materia_id = request.POST.get('materia', None)
        
        periodo = PeriodoAcademico.objects.get(id=periodo_id)
        
        # Filtrar cursos según criterios
        cursos = Curso.objects.filter(periodo=periodo)
        if programa_id:
            cursos = cursos.filter(materia__programa_id=programa_id)
        if materia_id:
            cursos = cursos.filter(materia_id=materia_id)
        
        if tipo_reporte == 'rendimiento_general':
            return generar_reporte_rendimiento_general(request, cursos, periodo, formato)
        elif tipo_reporte == 'estudiantes_riesgo':
            return generar_reporte_estudiantes_riesgo(request, periodo, formato)
        elif tipo_reporte == 'notas_por_materia':
            if materia_id:
                return generar_reporte_notas_materia(request, materia_id, periodo, formato)
    
    periodos = PeriodoAcademico.objects.all()
    programas = Programa.objects.filter(activo=True)
    materias = Materia.objects.all()
    
    context = {
        'periodos': periodos,
        'programas': programas,
        'materias': materias,
    }
    
    return render(request, 'administrador/generar_reporte.html', context)

def generar_reporte_rendimiento_general(request, cursos, periodo, formato):
    """Reporte de rendimiento académico general"""
    
    if formato == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo = Paragraph(f"<b>REPORTE DE RENDIMIENTO ACADÉMICO GENERAL</b>", styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información del reporte
        info = Paragraph(f"""
            <b>Periodo:</b> {periodo.nombre}<br/>
            <b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
            <b>Generado por:</b> {request.user.get_full_name()}
        """, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de datos
        data = [['Curso', 'Grupo', 'Profesor', 'Inscritos', 'Promedio', 'Aprobados']]
        
        total_estudiantes = 0
        total_aprobados = 0
        promedios_generales = []
        
        for curso in cursos:
            inscripciones = curso.inscripciones.all()
            total_inscritos = inscripciones.count()
            total_estudiantes += total_inscritos
            
            promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
            promedio_curso = sum(promedios) / len(promedios) if promedios else 0.0
            
            aprobados = sum(1 for p in promedios if p >= 3.0)
            total_aprobados += aprobados
            
            if promedio_curso > 0:
                promedios_generales.append(promedio_curso)
            
            data.append([
                f"{curso.materia.codigo}",
                curso.grupo,
                curso.profesor.usuario.last_name,
                str(total_inscritos),
                f"{promedio_curso:.2f}",
                f"{aprobados}/{total_inscritos}"
            ])
        
        table = Table(data, colWidths=[1.5*inch, 0.7*inch, 1.3*inch, 0.8*inch, 0.8*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Resumen estadístico
        promedio_institucional = sum(promedios_generales) / len(promedios_generales) if promedios_generales else 0.0
        tasa_aprobacion = (total_aprobados / total_estudiantes * 100) if total_estudiantes > 0 else 0.0
        
        elements.append(Spacer(1, 0.3*inch))
        resumen = Paragraph(f"""
            <b>RESUMEN ESTADÍSTICO</b><br/>
            Total de Estudiantes: {total_estudiantes}<br/>
            Total de Aprobados: {total_aprobados}<br/>
            Tasa de Aprobación: {tasa_aprobacion:.1f}%<br/>
            Promedio Institucional: <b>{promedio_institucional:.2f}</b>
        """, styles['Normal'])
        elements.append(resumen)
        
        doc.build(elements)
        buffer.seek(0)
        
        registrar_actividad(request, 'consultar', 'Reporte', periodo.id, 'Generación de reporte PDF')
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_rendimiento_{periodo.nombre}.pdf"'
        return response
    
    elif formato == 'excel':
        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rendimiento Académico"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        center_aligned = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['Curso', 'Grupo', 'Profesor', 'Inscritos', 'Promedio', 'Aprobados', 'Reprobados']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
        
        # Datos
        row = 2
        for curso in cursos:
            inscripciones = curso.inscripciones.all()
            total_inscritos = inscripciones.count()
            
            promedios = [insc.calcular_promedio() for insc in inscripciones if insc.calcular_promedio() is not None]
            promedio_curso = sum(promedios) / len(promedios) if promedios else 0.0
            
            aprobados = sum(1 for p in promedios if p >= 3.0)
            reprobados = len(promedios) - aprobados
            
            ws.cell(row=row, column=1).value = f"{curso.materia.codigo} - {curso.materia.nombre}"
            ws.cell(row=row, column=2).value = curso.grupo
            ws.cell(row=row, column=3).value = curso.profesor.usuario.get_full_name()
            ws.cell(row=row, column=4).value = total_inscritos
            ws.cell(row=row, column=5).value = round(promedio_curso, 2)
            ws.cell(row=row, column=6).value = aprobados
            ws.cell(row=row, column=7).value = reprobados
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = center_aligned
            
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        registrar_actividad(request, 'consultar', 'Reporte', periodo.id, 'Generación de reporte Excel')
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_rendimiento_{periodo.nombre}.xlsx"'
        return response

def generar_reporte_estudiantes_riesgo(request, periodo, formato):
    """Reporte de estudiantes en riesgo académico"""
    estudiantes_riesgo = []
    
    inscripciones = InscripcionCurso.objects.filter(curso__periodo=periodo)
    
    for insc in inscripciones:
        promedio = insc.calcular_promedio()
        if promedio and promedio < 3.0:
            estudiantes_riesgo.append({
                'estudiante': insc.estudiante,
                'curso': insc.curso,
                'promedio': promedio,
            })
    
    if formato == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        titulo = Paragraph("<b>REPORTE DE ESTUDIANTES EN RIESGO ACADÉMICO</b>", styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 0.2*inch))
        
        info = Paragraph(f"""
            <b>Periodo:</b> {periodo.nombre}<br/>
            <b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y')}<br/>
            <b>Total en riesgo:</b> {len(estudiantes_riesgo)} estudiantes
        """, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))
        
        data = [['Código', 'Estudiante', 'Materia', 'Promedio']]
        
        for item in estudiantes_riesgo:
            data.append([
                item['estudiante'].codigo_estudiantil,
                item['estudiante'].usuario.get_full_name(),
                item['curso'].materia.nombre,
                f"{item['promedio']:.2f}"
            ])
        
        table = Table(data, colWidths=[1.2*inch, 2*inch, 2.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C62828')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="estudiantes_riesgo_{periodo.nombre}.pdf"'
        return response

def generar_reporte_notas_materia(request, materia_id, periodo, formato):
    """Reporte detallado de notas por materia"""
    materia = Materia.objects.get(id=materia_id)
    cursos = Curso.objects.filter(materia=materia, periodo=periodo)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    titulo = Paragraph(f"<b>REPORTE DE NOTAS - {materia.nombre}</b>", styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 0.2*inch))
    
    for curso in cursos:
        elements.append(Paragraph(f"<b>Grupo: {curso.grupo} - Profesor: {curso.profesor.usuario.get_full_name()}</b>", styles['Heading2']))
        elements.append(Spacer(1, 0.1*inch))
        
        inscripciones = curso.inscripciones.all()
        data = [['Código', 'Estudiante', 'Promedio', 'Estado']]
        
        for insc in inscripciones:
            promedio = insc.calcular_promedio()
            data.append([
                insc.estudiante.codigo_estudiantil,
                insc.estudiante.usuario.get_full_name(),
                f"{promedio:.2f}" if promedio else "N/A",
                insc.estado_aprobacion()
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="notas_{materia.codigo}_{periodo.nombre}.pdf"'
    return response


# ==================== NOTIFICACIONES ====================

@login_required
def todas_notificaciones(request):
    """Ver todas las notificaciones del usuario"""
    notificaciones = request.user.notificaciones.all().order_by('-fecha_creacion')
    
    # Paginación
    paginator = Paginator(notificaciones, 20)
    page_number = request.GET.get('page')
    notificaciones_page = paginator.get_page(page_number)
    
    context = {
        'notificaciones': notificaciones_page,
    }
    
    return render(request, 'notificaciones.html', context)

@login_required
@require_http_methods(["POST"])
def marcar_notificacion_leida(request, notificacion_id):
    """Marcar notificación como leída (AJAX)"""
    notificacion = get_object_or_404(Notificacion, id=notificacion_id, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    
    return JsonResponse({'success': True})

@login_required
@require_http_methods(["POST"])
def marcar_todas_leidas(request):
    """Marcar todas las notificaciones como leídas (AJAX)"""
    request.user.notificaciones.filter(leida=False).update(leida=True)
    return JsonResponse({'success': True, 'message': 'Todas las notificaciones marcadas como leídas'})


# ==================== VISTAS MODALES/AJAX ====================

@login_required
def obtener_calificaciones_estudiante(request, inscripcion_id):
    """Obtener calificaciones de un estudiante en formato JSON (para modal)"""
    inscripcion = get_object_or_404(InscripcionCurso, id=inscripcion_id)
    
    # Verificar permisos
    if request.user.rol == 'estudiante':
        if inscripcion.estudiante.usuario != request.user:
            return JsonResponse({'error': 'Sin permisos'}, status=403)
    elif request.user.rol == 'profesor':
        if inscripcion.curso.profesor.usuario != request.user:
            return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    calificaciones = inscripcion.calificaciones.all()
    promedio = inscripcion.calcular_promedio()
    
    data = {
        'estudiante': inscripcion.estudiante.usuario.get_full_name(),
        'curso': inscripcion.curso.materia.nombre,
        'promedio': promedio,
        'estado': inscripcion.estado_aprobacion(),
        'calificaciones': [
            {
                'id': cal.id,
                'tipo': cal.tipo_evaluacion.nombre,
                'nota': float(cal.nota),
                'observaciones': cal.observaciones,
                'fecha': cal.fecha_registro.strftime('%d/%m/%Y %H:%M'),
            }
            for cal in calificaciones
        ]
    }
    
    return JsonResponse(data)

@login_required
@user_passes_test(es_administrador)
def estadisticas_dashboard(request):
    """Obtener estadísticas para el dashboard del admin (AJAX)"""
    periodo_id = request.GET.get('periodo')
    periodo = PeriodoAcademico.objects.get(id=periodo_id) if periodo_id else PeriodoAcademico.objects.filter(activo=True).first()
    
    cursos = Curso.objects.filter(periodo=periodo)
    
    # Calcular estadísticas
    total_cursos = cursos.count()
    total_inscripciones = InscripcionCurso.objects.filter(curso__periodo=periodo).count()
    
    promedios = []
    aprobados = 0
    reprobados = 0
    
    for curso in cursos:
        inscripciones = curso.inscripciones.all()
        for insc in inscripciones:
            promedio = insc.calcular_promedio()
            if promedio:
                promedios.append(promedio)
                if promedio >= 3.0:
                    aprobados += 1
                else:
                    reprobados += 1
    
    promedio_institucional = sum(promedios) / len(promedios) if promedios else 0.0
    
    data = {
        'total_cursos': total_cursos,
        'total_inscripciones': total_inscripciones,
        'promedio_institucional': round(promedio_institucional, 2),
        'aprobados': aprobados,
        'reprobados': reprobados,
        'tasa_aprobacion': round((aprobados / len(promedios) * 100) if promedios else 0, 1),
    }
    
    return JsonResponse(data)

@login_required
@require_http_methods(["POST"])
def validar_nota(request):
    """Validar formato de nota antes de guardar (AJAX)"""
    try:
        nota = float(request.POST.get('nota'))
        if 0.0 <= nota <= 5.0:
            return JsonResponse({'valid': True, 'nota': nota})
        else:
            return JsonResponse({'valid': False, 'error': 'La nota debe estar entre 0.0 y 5.0'})
    except (ValueError, TypeError):
        return JsonResponse({'valid': False, 'error': 'Formato de nota inválido'})


# ==================== BÚSQUEDA GLOBAL ====================

@login_required
def busqueda_global(request):
    """Búsqueda global en el sistema"""
    query = request.GET.get('q', '')
    
    if len(query) < 3:
        return JsonResponse({'results': [], 'message': 'Ingrese al menos 3 caracteres'})
    
    results = []
    
    if request.user.rol == 'profesor':
        # Buscar estudiantes en sus cursos
        cursos = request.user.perfil_profesor.cursos.all()
        inscripciones = InscripcionCurso.objects.filter(curso__in=cursos)
        inscripciones = inscripciones.filter(
            Q(estudiante__usuario__first_name__icontains=query) |
            Q(estudiante__usuario__last_name__icontains=query) |
            Q(estudiante__codigo_estudiantil__icontains=query)
        )[:10]
        
        for insc in inscripciones:
            results.append({
                'tipo': 'estudiante',
                'nombre': insc.estudiante.usuario.get_full_name(),
                'codigo': insc.estudiante.codigo_estudiantil,
                'url': f'/profesor/calificacion/{insc.id}/'
            })
    
    elif request.user.rol == 'administrador':
        # Buscar cursos, estudiantes y profesores
        cursos = Curso.objects.filter(
            Q(materia__nombre__icontains=query) |
            Q(materia__codigo__icontains=query)
        )[:5]
        
        for curso in cursos:
            results.append({
                'tipo': 'curso',
                'nombre': f"{curso.materia.codigo} - {curso.materia.nombre}",
                'grupo': curso.grupo,
                'url': f'/administrador/cursos/'
            })
        
        estudiantes = Estudiante.objects.filter(
            Q(usuario__first_name__icontains=query) |
            Q(usuario__last_name__icontains=query) |
            Q(codigo_estudiantil__icontains=query)
        )[:5]
        
        for est in estudiantes:
            results.append({
                'tipo': 'estudiante',
                'nombre': est.usuario.get_full_name(),
                'codigo': est.codigo_estudiantil,
                'programa': est.programa.nombre,
            })
    
    return JsonResponse({'results': results})


# ==================== EXPORTAR DATOS ====================

@login_required
@user_passes_test(es_estudiante)
def exportar_historial_notas(request):
    """Exportar historial completo de notas del estudiante"""
    estudiante = request.user.perfil_estudiante
    inscripciones = estudiante.inscripciones.all().order_by('-curso__periodo__fecha_inicio')
    
    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Notas"
    
    # Información del estudiante
    ws['A1'] = "HISTORIAL ACADÉMICO"
    ws['A2'] = f"Estudiante: {estudiante.usuario.get_full_name()}"
    ws['A3'] = f"Código: {estudiante.codigo_estudiantil}"
    ws['A4'] = f"Programa: {estudiante.programa.nombre}"
    
    # Encabezados
    headers = ['Periodo', 'Código', 'Materia', 'Créditos', 'Promedio', 'Estado']
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col).value = header
        ws.cell(row=6, column=col).font = Font(bold=True)
    
    # Datos
    row = 7
    for insc in inscripciones:
        promedio = insc.calcular_promedio()
        ws.cell(row=row, column=1).value = insc.curso.periodo.nombre
        ws.cell(row=row, column=2).value = insc.curso.materia.codigo
        ws.cell(row=row, column=3).value = insc.curso.materia.nombre
        ws.cell(row=row, column=4).value = insc.curso.materia.creditos
        ws.cell(row=row, column=5).value = round(promedio, 2) if promedio else "N/A"
        ws.cell(row=row, column=6).value = insc.estado_aprobacion()
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="historial_notas_{estudiante.codigo_estudiantil}.xlsx"'
    return response

@login_required
@user_passes_test(es_estudiante) 
def descargar_boletin_periodo(request, periodo_id):
    """Descargar boletín de notas en PDF para un periodo específico"""
    estudiante = request.user.perfil_estudiante
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id)
    
    inscripciones = estudiante.inscripciones.filter(curso__periodo=periodo)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo = Paragraph(f"<b>BOLETÍN DE NOTAS - {periodo.nombre}</b>", styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del estudiante
    info = Paragraph(f"""
        <b>Estudiante:</b> {estudiante.usuario.get_full_name()}<br/>
        <b>Código:</b> {estudiante.codigo_estudiantil}<br/>
        <b>Programa:</b> {estudiante.programa.nombre}<br/>
        <b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}
    """, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de calificaciones
    data = [['Código', 'Materia', 'Créditos', 'Promedio', 'Estado']]
    total_creditos = 0
    promedios_list = []
    
    for insc in inscripciones:
        promedio = insc.calcular_promedio()
        estado = insc.estado_aprobacion()
        creditos = insc.curso.materia.creditos
        total_creditos += creditos
        
        data.append([
            insc.curso.materia.codigo,
            insc.curso.materia.nombre,
            str(creditos),
            f"{promedio:.2f}" if promedio else "N/A",
            estado
        ])
        
        if promedio is not None:
            promedios_list.append(promedio)